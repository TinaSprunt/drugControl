#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
遍历 drugFoundPeopleRes/ 下的所有 .xlsx 表格文件，按身份证号拆分每一行。

- 扫描每行所有单元格提取身份证号，不依赖表头列名定位
- 每个源表格建立同名子文件夹，输出文件存入对应文件夹
- 每个表格单独生成 {表格名}-CountLog.xlsx 统计日志
- 详细处理日志输出至 drugFoundPeople.txt
- 无身份证号而跳过的行按原列结构写入各表子文件夹下的 {表格名}-error.xlsx

用法:
    python splitByIdCard.py
"""

import os
import re
import logging
import time
from collections import defaultdict

import openpyxl
from openpyxl.utils import get_column_letter

# ------------------ 可配置部分 ------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
RES_DIR = os.path.join(BASE_DIR, "drugFoundPeopleRes")             # 源文件目录
LOG_FILE = os.path.join(BASE_DIR, "drugFoundPeople.txt")           # 详细处理日志

# 身份证号正则：18位数字，或17位数字+末位X/x
ID_CARD_RE = re.compile(r'^(\d{17}[\dXx])$')

# 输出文件名匹配正则（避免将已生成的输出文件当作源文件处理）
OUTPUT_FILE_RE = re.compile(r'-\d{17}[\dXx]\.xlsx$')

# ------------------ 日志配置 ------------------
def setup_logging():
    """配置双通道日志：控制台 + 文件"""
    logger = logging.getLogger("SplitByIdCard")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()

    formatter = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s",
                                  datefmt="%Y-%m-%d %H:%M:%S")

    # 文件处理器
    fh = logging.FileHandler(LOG_FILE, mode='w', encoding='utf-8')
    fh.setLevel(logging.INFO)
    fh.setFormatter(formatter)

    # 控制台处理器
    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)
    ch.setFormatter(formatter)

    logger.addHandler(fh)
    logger.addHandler(ch)
    return logger


logger = setup_logging()


# ------------------ 工具函数 ------------------
def extract_id_card(value):
    """从单元格值中提取身份证号，提取失败返回 None。
    兼容前后空格、单引号(')、双引号(\")等边界字符。"""
    if value is None:
        return None
    # 转为字符串，去除单元格内的换行符（\r, \n），避免干扰正则匹配
    s = str(value).replace('\r', '').replace('\n', '')
    # 去除前后空格及引号包裹（如 '522601196807300814' 或 "522601196807300814"）
    s = s.strip().strip("'\"")
    # 精确匹配：整个字符串恰好是身份证号
    m = ID_CARD_RE.match(s)
    if m:
        return m.group(1).upper()
    # 兜底：从更复杂的字符串中搜索提取
    m = ID_CARD_RE.search(s)
    if m:
        return m.group(1).upper()
    return None


def is_source_file(filename: str) -> bool:
    """
    判断是否为需要处理的源文件。
    排除：临时文件、统计日志、已生成的输出文件、error.xlsx。
    """
    if not filename.lower().endswith('.xlsx'):
        return False
    if filename.startswith('~$'):
        return False
    if filename in ("drugFoundPeopleCountLog.xlsx", "error.xlsx"):
        return False
    if OUTPUT_FILE_RE.search(filename):
        return False
    if re.search(r'-CountLog\.xlsx$', filename):
        return False
    return True


def read_source_file(src_path: str, base_name: str):
    """
    读取单个源文件，按身份证号分组行数据。
    - 遍历每行所有单元格，用正则匹配身份证号，不依赖特定列
    - 同时收集无法匹配身份证号的跳过行

    返回:
        id_groups:  dict, {id_card: [row_data_list]}
        headers:    list, 表头列表
        skipped_rows: list of (row_number, row_data), 跳过的行
        total, matched, skipped: int 统计数
    """
    logger.info("读取: %s.xlsx", base_name)

    wb = openpyxl.load_workbook(src_path, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows()

    # ---- 读取表头 ----
    try:
        header_cells = next(rows_iter)
        headers = [cell.value for cell in header_cells]
    except StopIteration:
        wb.close()
        return {}, headers, [], 0, 0, 0

    # ---- 逐行扫描 ----
    id_groups = defaultdict(list)
    skipped_rows = []
    total = 0
    skipped = 0

    for row_cells in rows_iter:
        total += 1
        row_data = [cell.value for cell in row_cells]

        # 扫描本行所有单元格，提取身份证号
        id_card = None
        for cell_value in row_data:
            id_card = extract_id_card(cell_value)
            if id_card is not None:
                break

        if id_card is None:
            skipped += 1
            skipped_rows.append((total, row_data))  # total 即数据行号（从1开始）
            continue

        id_groups[id_card].append(row_data)

    wb.close()

    matched = sum(len(v) for v in id_groups.values())
    logger.info("  总行数: %d, 有效: %d, 跳过(无身份证号): %d", total, matched, skipped)
    return id_groups, headers, skipped_rows, total, matched, skipped


def write_output_file(output_path: str, headers: list, rows: list):
    """使用 openpyxl 创建工作簿并写入表头 + 数据行"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(output_path)
    wb.close()


def count_data_rows(file_path: str) -> int:
    """统计 .xlsx 文件中的有效数据行数（不含表头）"""
    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb.active
    row_count = max(ws.max_row - 1, 0) if ws.max_row else 0
    wb.close()
    return row_count


def write_count_log(output_dir: str, base_name: str, file_row_pairs: list, total_rows: int):
    """
    写入单表格的统计日志文件。

    参数:
        output_dir:     该表格的输出子文件夹
        base_name:      表格名（如 "羟考酮"）
        file_row_pairs: list of (filename, row_count)
        total_rows:     该表格的数据总行数
    """
    log_path = os.path.join(output_dir, f"{base_name}-CountLog.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "统计"
    ws.append(["文件名", "数据行数"])

    for fname, row_cnt in file_row_pairs:
        ws.append([fname, row_cnt])

    ws.append([])
    ws.append(["合计", total_rows])

    ws.column_dimensions[get_column_letter(1)].width = 50
    ws.column_dimensions[get_column_letter(2)].width = 15

    wb.save(log_path)
    wb.close()
    logger.info("  统计日志: %s (%d 个文件, %d 行)", log_path, len(file_row_pairs), total_rows)


def write_error_file(output_dir: str, base_name: str, headers: list, skipped_rows: list):
    """
    将单个表格中跳过的行写入 error.xlsx，每列独立占据单元格，保持数据结构完整。

    参数:
        output_dir:   该表格的输出子文件夹
        base_name:    表格名（如 "羟考酮"）
        headers:      源文件的表头列表
        skipped_rows: list of (row_number, row_data_list), 跳过的行
    """
    if not skipped_rows:
        logger.info("  无跳过行，不生成 error.xlsx")
        return

    error_path = os.path.join(output_dir, f"{base_name}-error.xlsx")

    # 裁剪尾部空列（去除所有尾部 None / 空字符串）
    def trim_tail(lst):
        cut = len(lst)
        while cut > 0 and (lst[cut - 1] is None or lst[cut - 1] == ''):
            cut -= 1
        return lst[:cut]

    trimmed_headers = trim_tail([str(h) if h is not None else "" for h in headers])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "跳过行"

    # 表头：来源表 + 数据行号 + 原表格的各列列名
    combined_headers = ["来源表", "数据行号"] + trimmed_headers
    ws.append(combined_headers)

    # 数据行：每列独立占据对应单元格，与表头对齐
    header_data_cols = len(trimmed_headers)
    for row_num, row_data in skipped_rows:
        # 裁剪后与原数据列数对齐
        trimmed_data = trim_tail([v if v is not None else "" for v in row_data])
        # 确保列数一致
        while len(trimmed_data) < header_data_cols:
            trimmed_data.append("")
        row_values = [base_name, row_num] + trimmed_data[:header_data_cols]
        ws.append(row_values)

    # 调整列宽
    ws.column_dimensions[get_column_letter(1)].width = 16
    ws.column_dimensions[get_column_letter(2)].width = 12

    wb.save(error_path)
    wb.close()
    logger.info("  跳过行记录: %s (%d 条, %d 列)", error_path, len(skipped_rows), len(combined_headers))


# ------------------ 主流程 ------------------
def main():
    start_time = time.time()
    logger.info("=" * 60)
    logger.info("按身份证号拆分 drugFoundPeopleRes 中的表格")
    logger.info("=" * 60)

    if not os.path.isdir(RES_DIR):
        logger.error("目录不存在: %s", RES_DIR)
        return

    # ---- 收集所有待处理的源文件 ----
    source_files = sorted([
        f for f in os.listdir(RES_DIR) if is_source_file(f)
    ])

    if not source_files:
        logger.info("没有找到需要处理的 .xlsx 源文件")
        return

    logger.info("找到 %d 个源文件: %s", len(source_files), source_files)

    # ---- 阶段: 逐表处理 ----
    total_files_created = 0
    total_skipped = 0

    for filename in source_files:
        src_path = os.path.join(RES_DIR, filename)
        base_name = os.path.splitext(filename)[0]

        logger.info("-" * 50)
        logger.info("处理: %s", filename)

        # 为该表格创建同名子文件夹
        table_output_dir = os.path.join(RES_DIR, base_name)
        os.makedirs(table_output_dir, exist_ok=True)

        # 读取源文件，提取身份证号
        id_groups, headers, skipped_rows, total_rows, matched_rows, skip_count = \
            read_source_file(src_path, base_name)

        # ---- 写入该表格的 error.xlsx（每列独立单元格） ----
        write_error_file(table_output_dir, base_name, headers, skipped_rows)
        total_skipped += skip_count

        if not id_groups:
            logger.info("  -> 无有效身份证号数据行")
            continue

        # ---- 写入按身份证号拆分的输出文件 ----
        created = 0
        skipped_existing = 0
        file_row_pairs = []

        for id_card, rows in id_groups.items():
            output_filename = f"{base_name}-{id_card}.xlsx"
            output_path = os.path.join(table_output_dir, output_filename)

            if os.path.exists(output_path):
                skipped_existing += 1
                row_count = count_data_rows(output_path)
                file_row_pairs.append((output_filename, row_count))
                continue

            write_output_file(output_path, headers, rows)
            file_row_pairs.append((output_filename, len(rows)))
            created += 1

        logger.info("  -> 新建 %d 个文件，跳过 %d 个已存在", created, skipped_existing)
        total_files_created += created

        # ---- 写入该表格的统计日志 ----
        write_count_log(table_output_dir, base_name, file_row_pairs, matched_rows)

    # ---- 汇总 ----
    elapsed = time.time() - start_time
    logger.info("=" * 60)
    logger.info("全部完成! 耗时 %.1f 秒", elapsed)
    logger.info("共处理 %d 个表格，新建 %d 个输出文件，跳过 %d 行",
                len(source_files), total_files_created, total_skipped)


if __name__ == "__main__":
    main()
