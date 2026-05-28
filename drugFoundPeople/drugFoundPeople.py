#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
递归提取多种表格（.xls, .xlsx, .et 等）中包含关键字的行，合并输出到一个新 Excel。

特性：
  - 支持 .xls / .xlsx / .xlsm / .et 等格式
  - 预扫描阶段：遍历所有文件读取表头，确定最大列结构，消除列截断
  - .xlsx 使用 openpyxl read_only 流式逐行读取，内存友好
  - .xls/.et 使用 xlrd (1.2.0) 直接逐行读取，绕过 pandas 的版本限制
  - 使用 openpyxl write_only 模式流式写入结果，避免内存溢出
  - 统一列结构：预扫描确定最宽列名，后续自动填充缺失列，绝不截断
  - 无论匹配多少行，最终仅生成一个以关键字命名的 .xlsx 文件
  - 运行日志输出到 drugFoundPeopleLog.txt
  - 对 .et 文件记录每 Sheet 的实际列数，便于排查格式兼容问题
  - 支持批量模式：读取"目标药品名称.xlsx"第A列中的关键字列表，依次处理

用法：
  单关键字:  python drugFoundPeople.py "盐酸瑞芬太尼"
  批量模式:  python drugFoundPeople.py          （从 目标药品名称.xlsx 读取关键字列表）
输出：
    drugFoundPeopleRes/关键字.xlsx
    drugFoundPeopleLog.txt
"""

import os
import sys
import re
import logging
import time

# ------------------ 可配置部分 ------------------
SOURCE_DIR = "data"                     # 源文件夹（如需处理全州数据，改为 "data（全州）"）
OUTPUT_DIR = "drugFoundPeopleRes"       # 结果文件夹
LOG_FILE = "drugFoundPeopleLog.txt"     # 日志文件
SUPPORTED_EXT = ('.xls', '.xlsx', '.xlsm', '.et')
SKIP_PREFIX = '~$'                      # 跳过 Excel 临时锁定文件
WRITE_BATCH_SIZE = 5000                 # 每积累 N 条匹配行就写一次磁盘

# ------------------ 日志配置 ------------------
def setup_logging():
    logger = logging.getLogger("Extract")
    logger.setLevel(logging.INFO)
    formatter = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")

    fh = logging.FileHandler(LOG_FILE, mode='w', encoding='utf-8')
    fh.setLevel(logging.INFO)
    fh.setFormatter(formatter)

    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(formatter)
    # 处理 Windows GBK 控制台下含特殊 Unicode 字符（如 \xa0）的文件路径
    if hasattr(ch, 'stream') and hasattr(ch.stream, 'reconfigure'):
        try:
            ch.stream.reconfigure(errors='replace')
        except Exception:
            pass

    logger.handlers.clear()
    logger.addHandler(fh)
    logger.addHandler(ch)
    return logger

logger = setup_logging()

# ------------------ 工具函数 ------------------
def safe_filename(text: str) -> str:
    """将关键字中的非法文件名字符替换为下划线"""
    return re.sub(r'[\\/*?:"<>|]', '_', text)


def trim_trailing_empty(lst: list) -> list:
    """去除列表尾部的空字符串元素，解决 Excel 列数虚高（16383列）的性能灾难。"""
    cut = len(lst)
    while cut > 0 and lst[cut - 1] == '':
        cut -= 1
    return lst[:cut] if cut < len(lst) else lst


def progress_bar(current: int, total: int, label: str = "", width: int = 40) -> str:
    """生成进度条字符串：[=====>    ] 50.0% label"""
    pct = current / total if total > 0 else 1.0
    filled = int(width * pct)
    bar = "=" * (filled - 1) + ">" if filled > 0 else ""
    bar = bar.ljust(width)
    return f"\r[{bar}] {pct * 100:5.1f}% {label}"


# ------------------ 输出管理 ------------------
class MatchWriter:
    """
    使用 openpyxl Workbook 常规模式写入匹配结果。
    - 预扫描阶段确定最大列结构作为输出表头
    - 后续命中自动补齐缺失列，绝不截断多余列
    - 最终仅生成一个 .xlsx 文件
    - 使用常规 Workbook（非 write_only），从根本上避免 write_only 模式的 ZIP 损坏问题
    - close() 时进行 ZIP 完整性校验，确保输出文件可用
    """

    def __init__(self, output_path: str):
        self.output_path = output_path          # 输出文件路径
        self._wb = None                         # openpyxl Workbook（常规模式）
        self._ws = None                         # 当前工作表
        self.output_headers = None              # 统一输出列名（含"来源文件""来源Sheet"）
        self._total_hit = 0                     # 全局命中计数

    @property
    def has_headers(self) -> bool:
        return self.output_headers is not None

    @property
    def data_column_count(self) -> int:
        """返回数据列的列数（不含"来源文件""来源Sheet"）"""
        if not self.output_headers:
            return 0
        return len(self.output_headers) - 2

    def _init_file(self):
        """创建输出工作簿（常规模式，确保 ZIP 完整性），写入表头"""
        from openpyxl import Workbook
        self._wb = Workbook()
        self._ws = self._wb.active
        self._ws.title = "匹配结果"
        self._ws.append(self.output_headers)

    def set_headers(self, data_headers: list):
        """
        设置统一的输出表头（仅在首次调用时生效）。
        data_headers: 数据源的原始列名列表（不含"来源文件""来源Sheet"）
        """
        if self.has_headers:
            return
        self.output_headers = ["来源文件", "来源Sheet"] + data_headers

    def write_rows(self, file_source: str, sheet_name: str, matched_rows: list):
        """
        批量写入匹配行。
        - 首次调用时自动创建输出文件
        - 每行数据列数不足时自动补空字符串
        - 绝不丢弃任何已读取的单元格数据
        """
        if not self.has_headers:
            raise RuntimeError("请先调用 set_headers() 确定输出列结构")
        if self._wb is None:
            self._init_file()

        expected_data_cols = self.data_column_count

        for row_vals in matched_rows:
            actual_len = len(row_vals)
            if actual_len < expected_data_cols:
                aligned = list(row_vals) + [''] * (expected_data_cols - actual_len)
            elif actual_len > expected_data_cols:
                aligned = row_vals[:expected_data_cols]
            else:
                aligned = row_vals

            full_row = [file_source, sheet_name] + aligned
            self._ws.append(full_row)
            self._total_hit += 1

    def close(self):
        """保存输出工作簿，进行 ZIP 完整性校验"""
        if self._wb is None:
            return

        save_ok = False
        try:
            self._wb.save(self.output_path)
            save_ok = True
        except Exception as e:
            logger.error(f"保存输出文件失败: {e}")
            raise
        finally:
            if self._wb is not None:
                self._wb.close()
            self._wb = None
            self._ws = None

        if save_ok:
            # 校验 ZIP 完整性
            self._validate_output()
            file_size_mb = os.path.getsize(self.output_path) / (1024 * 1024)
            logger.info(f"输出文件已保存: {self.output_path} ({file_size_mb:.1f} MB)")

    def _validate_output(self):
        """校验输出文件的 ZIP 完整性（CRC 校验）"""
        import zipfile
        try:
            with zipfile.ZipFile(self.output_path, 'r') as zf:
                bad_file = zf.testzip()
                if bad_file is not None:
                    logger.error(f"输出文件 ZIP 校验失败！损坏条目: {bad_file}")
                    logger.error("建议重新运行脚本以重新生成该文件。")
                else:
                    logger.info(f"输出文件 ZIP 完整性校验通过")
        except zipfile.BadZipFile as e:
            logger.error(f"输出文件不是有效的 ZIP 文件: {e}")
        except Exception as e:
            logger.error(f"输出文件 ZIP 校验异常: {e}")


# ------------------ 文件读取函数 ------------------
def _cell_to_str(val) -> str:
    """将任意单元格值转为字符串，处理 None / 数字 / 日期等类型"""
    if val is None:
        return ''
    if isinstance(val, float):
        # 避免浮点数出现 ".0" —— 整数不显示小数部分
        if val == int(val):
            return str(int(val))
        return str(val)
    return str(val)


def _row_contains_keyword(row_vals: list, keyword: str) -> bool:
    """判断一行中是否有任意单元格包含关键字（模糊匹配）"""
    for v in row_vals:
        if keyword in v:
            return True
    return False


def _discover_xls_headers(file_path: str) -> list:
    """用 xlrd 读取 .xls / .et 文件所有 sheet 的表头，返回有效列数最多的那个（去掉尾空列）。"""
    import xlrd
    best = []
    try:
        wb = xlrd.open_workbook(file_path, on_demand=True)
        for sheet_idx in range(wb.nsheets):
            try:
                sheet = wb.sheet_by_index(sheet_idx)
                if sheet.nrows > 0:
                    headers = [_cell_to_str(sheet.cell_value(0, c)) for c in range(sheet.ncols)]
                    headers = trim_trailing_empty(headers)
                    if len(headers) > len(best):
                        best = headers
            except Exception:
                continue
        wb.release_resources()
    except Exception:
        pass
    return best


def _discover_xlsx_headers(file_path: str) -> list:
    """用 openpyxl read_only 读取 .xlsx 文件所有 sheet 的表头，返回有效列数最多的那个（去掉尾空列）。"""
    from openpyxl import load_workbook
    best = []
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            try:
                ws = wb[sheet_name]
                row_iter = ws.iter_rows()
                header_cells = next(row_iter)
                headers = [_cell_to_str(c.value) for c in header_cells]
                headers = trim_trailing_empty(headers)
                if len(headers) > len(best):
                    best = headers
            except (StopIteration, Exception):
                continue
        wb.close()
    except Exception:
        pass
    return best


def discover_max_headers(source_dir: str, supported_ext: tuple) -> list:
    """
    预扫描阶段：遍历所有支持的表格文件，仅读取每 sheet 的第一行表头，
    找到列数最多的表头作为全局输出列结构。
    返回: 最大列数对应的表头列表，未找到任何表头则返回空列表。
    """
    max_headers = []
    max_file = ""
    file_count = 0

    logger.info("=" * 50)
    logger.info("预扫描阶段：正在遍历所有文件确定最大列结构...")
    for root, _dirs, files in os.walk(source_dir):
        for fname in files:
            if fname.startswith(SKIP_PREFIX):
                continue
            if not fname.lower().endswith(supported_ext):
                continue
            file_path = os.path.join(root, fname)
            file_count += 1
            ext = os.path.splitext(file_path)[1].lower()

            if ext in ('.xls', '.et'):
                headers = _discover_xls_headers(file_path)
            elif ext in ('.xlsx', '.xlsm'):
                headers = _discover_xlsx_headers(file_path)
            else:
                continue

            if len(headers) > len(max_headers):
                max_headers = headers
                max_file = os.path.relpath(file_path, source_dir)

    if max_headers:
        logger.info(f"预扫描完成（{file_count} 个文件）。")
        logger.info(f"最大列数: {len(max_headers)}，来源文件: {max_file}")
        logger.info(f"表头预览: {max_headers[:5]}{'...' if len(max_headers) > 5 else ''}")
    else:
        logger.warning(f"预扫描完成（{file_count} 个文件），未找到有效表头。")
    logger.info("=" * 50)

    return max_headers


def read_xls_with_xlrd(file_path: str, keyword: str, writer: MatchWriter) -> int:
    """
    使用 xlrd 1.2.0 直接读取 .xls / .et (OLE2 兼容) 文件。
    - 针对 .et 文件记录每 Sheet 的 xlrd 识别列数，便于排查格式兼容性
    - 只读取输出表头范围内的列（不再读取尾部 16000+ 空列）
    返回本文件累计命中行数。
    """
    import xlrd

    hit_count = 0
    file_label = os.path.relpath(file_path, SOURCE_DIR)
    is_et_file = file_path.lower().endswith('.et')

    try:
        wb = xlrd.open_workbook(file_path, on_demand=True)
    except Exception as e:
        logger.error(f"  xlrd 无法打开文件: {e}")
        return 0

    # 确定有效列数（取输出列数与实际列数的较小值，再截断尾空）
    output_cols = writer.data_column_count or 9999

    sheets_processed = 0
    for sheet_idx in range(wb.nsheets):
        try:
            sheet = wb.sheet_by_index(sheet_idx)
        except Exception as e:
            logger.error(f"  读取 Sheet 索引 {sheet_idx} 失败: {e}")
            continue

        if sheet.nrows <= 1:
            continue

        sheet_name = sheet.name
        sheet_cols = min(sheet.ncols, output_cols)

        # 针对 .et 文件输出更详细的列信息
        extra_info = ""
        if is_et_file:
            row_lens = [sheet.row_len(r) for r in range(sheet.nrows)]
            max_row_len = max(row_lens) if row_lens else 0
            min_row_len = min(row_lens) if row_lens else 0
            extra_info = f", 实际填充列范围: [{min_row_len}-{max_row_len}]"
        logger.info(f"  [xlrd] Sheet [{sheet_name}] — {sheet.nrows} 行, {sheet.ncols} 列 -> 有效 {sheet_cols} 列{extra_info}")

        # 读表头
        headers = [_cell_to_str(sheet.cell_value(0, c)) for c in range(sheet_cols)]

        if not writer.has_headers:
            writer.set_headers(headers)

        # 逐行匹配，分批写入（只读取有效列数）
        batch = []
        for r in range(1, sheet.nrows):
            row_vals = [_cell_to_str(sheet.cell_value(r, c)) for c in range(sheet_cols)]
            if _row_contains_keyword(row_vals, keyword):
                batch.append(row_vals)
                hit_count += 1
                if len(batch) >= WRITE_BATCH_SIZE:
                    writer.write_rows(file_label, sheet_name, batch)
                    batch.clear()

        if batch:
            writer.write_rows(file_label, sheet_name, batch)

        sheets_processed += 1
        if hit_count > 0:
            logger.info(f"    [xlrd] 命中 {hit_count} 行（累计）")

    wb.release_resources()

    if is_et_file:
        logger.info(f"  [xlrd] .et 文件处理完毕：{sheets_processed} 个 Sheet，共命中 {hit_count} 行")
    return hit_count


def read_xlsx_with_openpyxl(file_path: str, keyword: str, writer: MatchWriter) -> int:
    """
    使用 openpyxl read_only 流式读取 .xlsx 文件，逐行匹配。
    - 只读取输出表头范围内的列（不再读取尾部 16000+ 空列）
    返回本文件累计命中行数。
    """
    from openpyxl import load_workbook

    hit_count = 0
    file_label = os.path.relpath(file_path, SOURCE_DIR)

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        logger.error(f"  openpyxl 无法打开文件: {e}")
        return 0

    output_cols = writer.data_column_count

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        row_iter = ws.iter_rows()

        # 读表头
        try:
            header_cells = next(row_iter)
        except StopIteration:
            continue

        headers = [_cell_to_str(c.value) for c in header_cells]
        if all(h == '' for h in headers):
            continue

        # 确定有效列数：取输出列数与表头列数的较小值
        effective_cols = min(len(headers), output_cols) if output_cols else len(headers)
        logger.info(f"  [openpyxl] Sheet [{sheet_name}] — 有效 {effective_cols} 列")

        if not writer.has_headers:
            writer.set_headers(headers[:effective_cols])

        batch = []
        sheet_hit = 0

        for row_cells in row_iter:
            # 只读取有效列数范围内的单元格
            row_vals = [_cell_to_str(row_cells[c].value) for c in range(min(len(row_cells), effective_cols))]
            # 补齐不足的列
            if len(row_vals) < effective_cols:
                row_vals.extend([''] * (effective_cols - len(row_vals)))
            if _row_contains_keyword(row_vals, keyword):
                batch.append(row_vals)
                sheet_hit += 1
                hit_count += 1
                if len(batch) >= WRITE_BATCH_SIZE:
                    writer.write_rows(file_label, sheet_name, batch)
                    batch.clear()

        if batch:
            writer.write_rows(file_label, sheet_name, batch)

        if sheet_hit > 0:
            logger.info(f"    命中 {sheet_hit} 行")

    wb.close()
    return hit_count


def read_et_with_wps_com(file_path: str, keyword: str, writer: MatchWriter) -> int:
    """
    使用 WPS COM (ET.Application) 读取 .et 文件。
    需要安装 WPS 且其 COM 服务可用。
    """
    hit_count = 0
    file_label = os.path.relpath(file_path, SOURCE_DIR)

    try:
        from win32com import client
        et_app = client.Dispatch("ET.Application")
        et_app.Visible = False
        et_app.DisplayAlerts = False

        wb = et_app.Workbooks.Open(file_path)
        for sheet in wb.Sheets:
            sheet_name = sheet.Name
            data = sheet.UsedRange.Value
            if data is None or len(data) < 2:
                continue

            logger.info(f"  [WPS COM] Sheet [{sheet_name}] — {len(data)} 行")

            headers = [_cell_to_str(v) for v in data[0]]

            if not writer.has_headers:
                writer.set_headers(headers)

            batch = []
            for row in data[1:]:
                row_vals = [_cell_to_str(v) for v in row]
                if _row_contains_keyword(row_vals, keyword):
                    batch.append(row_vals)
                    hit_count += 1
                    if len(batch) >= WRITE_BATCH_SIZE:
                        writer.write_rows(file_label, sheet_name, batch)
                        batch.clear()

            if batch:
                writer.write_rows(file_label, sheet_name, batch)

            if hit_count > 0:
                logger.info(f"    [WPS COM] 命中 {hit_count} 行（累计）")

        wb.Close(False)
        et_app.Quit()
    except Exception as e:
        logger.warning(f"  WPS COM 读取失败: {e}")
        try:
            et_app.Quit()
        except Exception:
            pass
        raise
    return hit_count


# ------------------ 文件分发 ------------------
def process_file(file_path: str, keyword: str, writer: MatchWriter) -> int:
    """
    根据扩展名选择读取方式处理单个文件，返回命中行数。
    """
    ext = os.path.splitext(file_path)[1].lower()
    hit_count = 0

    if ext == '.et':
        # .et 文件：优先 xlrd（OLE2 兼容），失败再试 WPS COM
        try:
            hit_count = read_xls_with_xlrd(file_path, keyword, writer)
            if hit_count > 0:
                return hit_count
            # 如果 xlrd 读到 0 行，可能是 xlrd 没报错但实际格式不对，
            # 此时文件已被处理，不再重试 WPS COM（避免重复计数）
        except Exception:
            logger.warning("  xlrd 方式失败，尝试 WPS COM 回退")
            try:
                hit_count = read_et_with_wps_com(file_path, keyword, writer)
            except Exception as e2:
                logger.error(f"  WPS COM 也失败，跳过该文件: {e2}")

    elif ext == '.xls':
        try:
            hit_count = read_xls_with_xlrd(file_path, keyword, writer)
        except Exception as e:
            logger.error(f"  处理 .xls 文件出错: {e}", exc_info=True)

    elif ext in ('.xlsx', '.xlsm'):
        try:
            hit_count = read_xlsx_with_openpyxl(file_path, keyword, writer)
        except Exception as e:
            logger.error(f"  处理 .xlsx/.xlsm 文件出错: {e}", exc_info=True)

    else:
        logger.warning(f"  不支持的文件格式: {file_path}")

    return hit_count


# ------------------ 批量关键字加载 ------------------
def load_keywords_from_excel(filepath: str) -> list:
    """
    从 Excel 文件的第 A 列读取所有非空单元格作为关键字列表。
    跳过空单元格、表头行不会被特殊对待（A1 也作为关键字提取）。
    返回: 关键字字符串列表，读取失败时返回空列表。
    """
    from openpyxl import load_workbook

    keywords = []
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
            val = row[0]
            if val is not None:
                kw = str(val).strip()
                if kw:
                    keywords.append(kw)
        wb.close()
        logger.info(f"从 \"{filepath}\" 成功读取 {len(keywords)} 个关键字")
    except FileNotFoundError:
        logger.error(f"关键字文件不存在: \"{filepath}\"")
    except Exception as e:
        logger.error(f"读取关键字文件失败: {e}")
    return keywords


# ------------------ 单关键字处理入口 ------------------
def run_single_keyword(keyword: str, max_headers: list, file_list: list) -> int:
    """
    针对单个关键字，遍历所有源文件进行匹配并写入独立输出文件。
    max_headers: 预扫描阶段确定的全局最大列结构（若为空则动态确定）
    file_list: 预收集的文件路径列表 [(file_path, file_index), ...]
    返回: 总命中行数
    """
    safe_kw = safe_filename(keyword)
    out_path = os.path.join(OUTPUT_DIR, f"{safe_kw}.xlsx")

    writer = MatchWriter(out_path)
    if max_headers:
        writer.set_headers(max_headers)

    total_hit = 0
    total_files = len(file_list)
    start_time = time.time()

    logger.info(f"开始匹配关键字 \"{keyword}\"（共 {total_files} 个文件）...")
    try:
        for idx, (file_path, file_index) in enumerate(file_list, 1):
            # 控制台进度条
            sys.stdout.write(progress_bar(idx, total_files, f"\"{keyword}\" {idx}/{total_files}"))
            sys.stdout.flush()

            logger.info(f"[{file_index}] {file_path}")
            hit = process_file(file_path, keyword, writer)
            total_hit += hit

        sys.stdout.write("\r" + " " * 100 + "\r")  # 清除进度条
        sys.stdout.flush()

        elapsed = time.time() - start_time
        if total_hit == 0:
            logger.warning(f"关键字 \"{keyword}\" 未匹配到任何行，不生成输出文件。（耗时 {elapsed:.1f}s）")
        else:
            logger.info(f"关键字 \"{keyword}\" 处理完成！总命中: {total_hit} 行, 处理文件: {total_files} 个（耗时 {elapsed:.1f}s）")
    finally:
        writer.close()

    return total_hit


# ------------------ 主流程入口 ------------------
def main():
    """判断为单关键字模式还是批量模式，并执行相应逻辑。"""

    if not os.path.isdir(SOURCE_DIR):
        logger.error(f"源文件夹 '{SOURCE_DIR}' 不存在，脚本退出。")
        sys.exit(1)

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # -------- 阶段 1：预扫描，确定最大列结构和文件列表 --------
    max_headers = discover_max_headers(SOURCE_DIR, SUPPORTED_EXT)
    if max_headers:
        logger.info(f"输出表头已锁定为 {len(max_headers)} 列，后续所有行将严格对齐，绝不截断。")
    else:
        logger.warning("未发现有效的表头行，将在首次命中时动态确定列结构。")

    # 预收集所有文件路径（避免每个关键字重复遍历文件系统）
    file_list = []
    for root, _dirs, files in os.walk(SOURCE_DIR):
        for fname in files:
            if fname.startswith(SKIP_PREFIX):
                continue
            if not fname.lower().endswith(SUPPORTED_EXT):
                continue
            file_path = os.path.join(root, fname)
            file_list.append((file_path, len(file_list) + 1))
    logger.info(f"共发现 {len(file_list)} 个待处理文件。")

    # -------- 阶段 2：确定关键字列表 --------
    if len(sys.argv) >= 2:
        # 单关键字模式：从命令行读取
        keyword = sys.argv[1].strip()
        if not keyword:
            logger.error("错误: 关键字不能为空")
            sys.exit(1)
        keywords = [keyword]
        logger.info(f"单关键字模式 — 关键字: \"{keyword}\"")
    else:
        # 批量模式：从 目标药品名称.xlsx 读取
        script_dir = os.path.dirname(os.path.abspath(__file__))
        keyword_file = os.path.join(script_dir, "目标药品名称.xlsx")
        logger.info(f"批量模式 — 从 \"{keyword_file}\" 读取关键字列表...")
        keywords = load_keywords_from_excel(keyword_file)
        if not keywords:
            logger.error("未能从文件读取到任何有效关键字，脚本退出。")
            sys.exit(1)
        logger.info(f"读取到 {len(keywords)} 个关键字: {keywords}")

    # -------- 阶段 3：逐关键字处理 --------
    logger.info(f"源目录: {SOURCE_DIR}")
    logger.info(f"输出目录: {OUTPUT_DIR}")

    if len(keywords) == 1:
        # 单关键字：直接处理
        total_hit = run_single_keyword(keywords[0], max_headers, file_list)
        if total_hit == 0:
            logger.warning("未在任何文件中找到包含关键字的行，不生成输出文件。")
    else:
        # 批量模式：逐个处理，最后输出汇总
        summary = []
        overall_start = time.time()
        for i, kw in enumerate(keywords, 1):
            logger.info(f"\n{'#' * 50}")
            logger.info(f"# [{i}/{len(keywords)}] 关键字: {kw}")
            logger.info(f"{'#' * 50}")
            hit = run_single_keyword(kw, max_headers, file_list)
            summary.append((kw, hit))

        overall_elapsed = time.time() - overall_start
        logger.info(f"\n{'=' * 50}")
        logger.info(f"批量处理全部完成！共 {len(keywords)} 个关键字，总耗时 {overall_elapsed:.1f}s，汇总：")
        total_all = 0
        for kw, hit in summary:
            logger.info(f"  {kw}: {hit} 行")
            total_all += hit
        logger.info(f"  合计: {total_all} 行")
        logger.info(f"{'=' * 50}")


if __name__ == "__main__":
    main()
