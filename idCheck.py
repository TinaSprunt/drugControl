import os
import re
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font
from python_calamine import load_workbook as calamine_load

# ------------------------------
# 1. 身份证号提取与校验
# ------------------------------

def extract_id_numbers(text):
    if not isinstance(text, str):
        text = str(text) if text is not None else ""
    pattern_18 = r'\b[1-9]\d{5}(?:19|20)\d{2}(?:0[1-9]|1[0-2])(?:0[1-9]|[12]\d|3[01])\d{3}[\dXx]\b'
    pattern_15 = r'\b[1-9]\d{5}\d{2}(?:0[1-9]|1[0-2])(?:0[1-9]|[12]\d|3[01])\d{3}\b'
    candidates = re.findall(f'{pattern_18}|{pattern_15}', text)
    ids = set()
    for c in candidates:
        if isinstance(c, tuple):
            c = ''.join(c)
        if is_valid_id(c):
            ids.add(c.upper())
    return ids

def is_valid_id(id_str):
    id_str = id_str.strip().upper()
    if len(id_str) == 15:
        try:
            datetime.strptime(id_str[6:12], '%y%m%d')
            return True
        except ValueError:
            return False
    elif len(id_str) == 18:
        try:
            datetime.strptime(id_str[6:14], '%Y%m%d')
        except ValueError:
            return False
        weights = [7, 9, 10, 5, 8, 4, 2, 1, 6, 3, 7, 9, 10, 5, 8, 4, 2]
        check_codes = '10X98765432'
        if not id_str[:17].isdigit():
            return False
        total = sum(int(id_str[i]) * weights[i] for i in range(17))
        if check_codes[total % 11] != id_str[17]:
            return False
        return True
    return False

def get_age(birth_date, today=None):
    if today is None:
        today = date.today()
    age = today.year - birth_date.year
    if (today.month, today.day) < (birth_date.month, birth_date.day):
        age -= 1
    return age

def is_minor(id_str, today=None):
    id_str = id_str.strip().upper()
    try:
        if len(id_str) == 18:
            birth = datetime.strptime(id_str[6:14], '%Y%m%d').date()
        else:
            birth = datetime.strptime(id_str[6:12], '%y%m%d').date()
        age = get_age(birth, today)
        return age < 18
    except (ValueError, IndexError):
        return False

# ------------------------------
# 2. 读取Excel（calamine引擎）
# ------------------------------

def read_excel_cells(filepath):
    ids = set()
    try:
        wb = calamine_load(filepath)
        for sheet_name in wb.sheet_names:
            sheet = wb.get_sheet_by_name(sheet_name)
            for row in sheet.to_python():
                for cell_value in row:
                    if cell_value is not None:
                        ids.update(extract_id_numbers(cell_value))
        wb.close()
    except Exception as e:
        print(f"  [错误] 读取失败: {e}")
    return ids

# ------------------------------
# 3. 主流程（递归 + 显式状态）
# ------------------------------

def main():
    data_dir = 'data'
    if not os.path.isdir(data_dir):
        print(f"致命错误：文件夹 '{data_dir}' 不存在！")
        input("按回车键退出...")
        return

    # 收集所有 Excel 文件
    excel_files = []
    print(f"正在递归扫描 {data_dir} 中的 Excel 文件...")
    for root, dirs, files in os.walk(data_dir):
        for f in files:
            if f.endswith(('.xlsx', '.xls', '.xlsm', '.xlsb')):
                full_path = os.path.join(root, f)
                excel_files.append(full_path)
                print(f"  找到: {full_path}")

    if not excel_files:
        print("未找到任何 Excel 文件，无法生成结果。")
        input("按回车键退出...")
        return

    all_ids = set()
    for filepath in excel_files:
        print(f"\n正在处理: {filepath}")
        ids = read_excel_cells(filepath)
        print(f"  该文件提取到 {len(ids)} 个有效身份证号")
        all_ids.update(ids)

    print(f"\n===== 汇总 =====")
    print(f"总共发现 {len(all_ids)} 个唯一的身份证号")

    # ---- 无论是否为空，都生成 newTarget.xlsx ----
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "身份证号"
        ws.append(["身份证号"])

        red_font = Font(color="FF0000")
        today = date.today()

        if all_ids:
            for id_num in sorted(all_ids):
                ws.append([id_num])
                cell = ws.cell(row=ws.max_row, column=1)
                if is_minor(id_num, today):
                    cell.font = red_font
            print(f"未成年身份证号已标记为红色（截至 {today}）")
        else:
            print("未提取到任何身份证号，将生成只有表头的空文件。")
            # 可以添加一行提示（可选）
            ws.append(["未发现身份证号"])

        output_file = "newTarget.xlsx"
        wb.save(output_file)
        print(f"\n✅ 文件已成功生成: {os.path.abspath(output_file)}")

    except Exception as e:
        print(f"\n❌ 保存文件时发生异常: {e}")
        print("可能原因：文件被占用、磁盘空间不足、权限不够等。")

    input("\n按回车键退出...")

if __name__ == "__main__":
    main()