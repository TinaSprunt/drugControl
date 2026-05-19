import os
import re
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font
from python_calamine import load_workbook as calamine_load
import olefile
import time

#  递归提取同级目录中data文件夹下的所有表格文件中的身份证号
#  2.0版本新增对wps表格 .et 文件的支持

# 固定工作目录为脚本所在目录
os.chdir(os.path.dirname(os.path.abspath(__file__)))

# ------------------------------
# 1. 身份证号提取与校验
# ------------------------------

def extract_id_numbers(text):
    if not isinstance(text, str):
        text = str(text) if text is not None else ""
    pattern_18 = r'\b[1-9]\d{5}(?:19|20)\d{2}(?:0[1-9]|1[0-2])(?:0[1-9]|[12]\d|3[01])\d{3}[\dXx]\b'
    pattern_15 = r'\b[1-9]\d{5}\d{2}(?:0[1-9]|1[0-2])(?:0[1-9]|[12]\d|3[01])\d{3}\b'
    candidates = re.findall(f'{pattern_18}|{pattern_15}', text)
    ids = set()
    for c in candidates:
        if isinstance(c, tuple):
            c = ''.join(c)
        if is_valid_id(c):
            ids.add(c.upper())
    return ids

def is_valid_id(id_str):
    id_str = id_str.strip().upper()
    if len(id_str) == 15:
        try:
            datetime.strptime(id_str[6:12], '%y%m%d')
            return True
        except ValueError:
            return False
    elif len(id_str) == 18:
        try:
            datetime.strptime(id_str[6:14], '%Y%m%d')
        except ValueError:
            return False
        weights = [7, 9, 10, 5, 8, 4, 2, 1, 6, 3, 7, 9, 10, 5, 8, 4, 2]
        check_codes = '10X98765432'
        if not id_str[:17].isdigit():
            return False
        total = sum(int(id_str[i]) * weights[i] for i in range(17))
        if check_codes[total % 11] != id_str[17]:
            return False
        return True
    return False

def get_age(birth_date, today=None):
    if today is None:
        today = date.today()
    age = today.year - birth_date.year
    if (today.month, today.day) < (birth_date.month, birth_date.day):
        age -= 1
    return age

def is_minor(id_str, today=None):
    id_str = id_str.strip().upper()
    try:
        if len(id_str) == 18:
            birth = datetime.strptime(id_str[6:14], '%Y%m%d').date()
        else:
            birth = datetime.strptime(id_str[6:12], '%y%m%d').date()
        age = get_age(birth, today)
        return age < 18
    except (ValueError, IndexError):
        return False

# ------------------------------
# 2. 文件读取引擎
# ------------------------------

def read_excel_cells_calamine(filepath):
    ids = set()
    try:
        wb = calamine_load(filepath)
        for sheet_name in wb.sheet_names:
            sheet = wb.get_sheet_by_name(sheet_name)
            for row in sheet.to_python():
                for cell_value in row:
                    if cell_value is not None:
                        ids.update(extract_id_numbers(cell_value))
        wb.close()
    except Exception as e:
        print(f"  [错误] calamine 读取失败: {e}")
    return ids

def read_et_by_wps(filepath):
    """通过 WPS COM 接口读取 .et 文件，管理员权限运行才能成功"""
    ids = set()
    try:
        import win32com.client
    except ImportError:
        print("  [提示] 需要安装 pywin32 才能调用 WPS，请执行: pip install pywin32")
        return ids

    # 尝试的 ProgID 列表（根据实际测试调整，常见的为前两个）
    prog_ids = ["ET.Application", "KET.Application", "WPS.ET.Application", "WPS.Application"]
    et = None
    for prog in prog_ids:
        try:
            et = win32com.client.Dispatch(prog)
            print(f"  [WPS] 成功使用 ProgID: {prog}")
            break
        except Exception:
            continue

    if et is None:
        print("  [错误] 未能创建 WPS 表格对象，请检查 WPS 安装，并确保以管理员身份运行脚本")
        return ids

    try:
        et.Visible = False
        et.DisplayAlerts = False
        abs_path = os.path.abspath(filepath)
        # 只读方式打开，避免修改原文件
        wb = et.Workbooks.Open(abs_path, ReadOnly=True, Format='et')
        time.sleep(1)  # 稍等文件加载

        for ws in wb.Worksheets:
            used = ws.UsedRange
            if used is None:
                continue
            rows = used.Rows.Count
            cols = used.Columns.Count
            for r in range(1, rows + 1):
                for c in range(1, cols + 1):
                    try:
                        val = ws.Cells(r, c).Value
                        if val is not None:
                            ids.update(extract_id_numbers(val))
                    except Exception:
                        continue
        wb.Close(False)
    except Exception as e:
        print(f"  [错误] WPS 读取失败: {e}")
        print("  请确认：1) 以管理员身份运行  2) .et 文件未损坏  3) WPS 完整安装")
    finally:
        try:
            et.Quit()
        except Exception:
            pass
    return ids

def read_et_by_olefile(filepath):
    """olefile 兜底方案，尝试从二进制流中提取身份证号文本"""
    ids = set()
    try:
        if not olefile.isOleFile(filepath):
            return ids
        ole = olefile.OleFileIO(filepath)
        # 收集所有流的内容
        all_texts = []
        for stream in ole.listdir():
            try:
                data = ole.openstream(stream).read()
                # 尝试多种常见编码
                for enc in ('utf-16-le', 'gbk', 'utf-8'):
                    try:
                        text = data.decode(enc, errors='ignore')
                        all_texts.append(text)
                        break
                    except:
                        continue
            except:
                continue
        ole.close()
        # 合并文本后提取身份证号
        combined = ' '.join(all_texts)
        ids = extract_id_numbers(combined)
    except Exception as e:
        print(f"  [警告] olefile 后备读取失败: {e}")
    return ids

def read_et_file(filepath):
    """.et 文件读取调度：先尝试 WPS COM，失败则用 olefile"""
    ids = read_et_by_wps(filepath)
    if not ids:
        print("  WPS 未提取到身份证号，尝试用 olefile 后备方案...")
        ids = read_et_by_olefile(filepath)
    return ids

def read_excel_cells(filepath):
    ext = os.path.splitext(filepath)[1].lower()
    if ext in ['.xlsx', '.xls', '.xlsm', '.xlsb']:
        return read_excel_cells_calamine(filepath)
    elif ext == '.et':
        return read_et_file(filepath)
    else:
        return set()

# ------------------------------
# 3. 主流程
# ------------------------------

def main():
    data_dir = 'data'
    if not os.path.isdir(data_dir):
        print(f"致命错误：文件夹 '{data_dir}' 不存在!")
        input("按回车键退出...")
        return

    supported_ext = ('.xlsx', '.xls', '.xlsm', '.xlsb', '.et')
    excel_files = []

    print(f"正在递归扫描 {data_dir} 中的表格文件...")
    for root, dirs, files in os.walk(data_dir):
        for f in files:
            if f.lower().endswith(supported_ext):
                full_path = os.path.join(root, f)
                excel_files.append(full_path)
                print(f"  找到: {full_path}")

    if not excel_files:
        print("未找到任何表格文件，无法生成结果。")
        input("按回车键退出...")
        return

    all_ids = set()
    for filepath in excel_files:
        print(f"\n正在处理: {filepath}")
        ids = read_excel_cells(filepath)
        print(f"  该文件提取到 {len(ids)} 个有效身份证号")
        all_ids.update(ids)

    print(f"\n===== 汇总 =====")
    print(f"总共发现 {len(all_ids)} 个唯一的身份证号")

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "身份证号"
        ws.append(["身份证号"])

        red_font = Font(color="FF0000")
        today = date.today()

        if all_ids:
            for id_num in sorted(all_ids):
                ws.append([id_num])
                cell = ws.cell(row=ws.max_row, column=1)
                if is_minor(id_num, today):
                    cell.font = red_font
            print(f"未成年身份证号已标记为红色（截至 {today}）")
        else:
            print("未提取到任何身份证号，将生成只有表头的空文件。")

        output_file = "newTarget.xlsx"
        wb.save(output_file)
        abs_path = os.path.abspath(output_file)
        print(f"\n✅ 文件已保存到: {abs_path}")
        if os.path.exists(abs_path):
            print(f"✅ 文件确认存在，大小: {os.path.getsize(abs_path)} 字节")
        else:
            print("❌ 文件未成功写入，请检查磁盘权限或杀毒软件。")

    except Exception as e:
        print(f"\n❌ 保存文件时发生异常: {e}")

    input("\n按回车键退出...")

if __name__ == "__main__":
    main()