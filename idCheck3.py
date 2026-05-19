import os
import re
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font
from python_calamine import load_workbook as calamine_load
import olefile
import time

# 固定工作目录为脚本所在文件夹
os.chdir(os.path.dirname(os.path.abspath(__file__)))

# ------------------------------
# 1. 身份证号提取与校验（改用无宽度边界）
# ------------------------------

def safe_str(val):
    """将单元格值转为字符串，避免科学计数法"""
    if val is None:
        return ""
    if isinstance(val, float):
        # 如果是整数值（如身份证号），转为 int 再 str
        if val == int(val) and not (val < 0 or val > 1e30):
            return str(int(val))
    if isinstance(val, (int, float)):
        return str(val)
    return str(val)

def extract_id_numbers(value):
    text = safe_str(value)
    if not text:
        return set()
    # 使用 (?<!\d) 和 (?!\d) 代替 \b，避免身份证号紧贴中文时漏掉
    pattern_18 = r'(?<!\d)[1-9]\d{5}(?:19|20)\d{2}(?:0[1-9]|1[0-2])(?:0[1-9]|[12]\d|3[01])\d{3}[\dXx](?!\d)'
    pattern_15 = r'(?<!\d)[1-9]\d{5}\d{2}(?:0[1-9]|1[0-2])(?:0[1-9]|[12]\d|3[01])\d{3}(?!\d)'
    candidates = re.findall(f'{pattern_18}|{pattern_15}', text)
    ids = set()
    for c in candidates:
        if isinstance(c, tuple):
            c = ''.join(c)
        if is_valid_id(c):
            ids.add(c.upper())
    return ids

def is_valid_id(id_str):
    id_str = id_str.strip().upper()
    if len(id_str) == 15:
        try:
            datetime.strptime(id_str[6:12], '%y%m%d')
            return True
        except ValueError:
            return False
    elif len(id_str) == 18:
        try:
            datetime.strptime(id_str[6:14], '%Y%m%d')
        except ValueError:
            return False
        weights = [7, 9, 10, 5, 8, 4, 2, 1, 6, 3, 7, 9, 10, 5, 8, 4, 2]
        check_codes = '10X98765432'
        if not id_str[:17].isdigit():
            return False
        total = sum(int(id_str[i]) * weights[i] for i in range(17))
        if check_codes[total % 11] != id_str[17]:
            return False
        return True
    return False

def get_age(birth_date, today=None):
    if today is None:
        today = date.today()
    age = today.year - birth_date.year
    if (today.month, today.day) < (birth_date.month, birth_date.day):
        age -= 1
    return age

def is_minor(id_str, today=None):
    id_str = id_str.strip().upper()
    try:
        if len(id_str) == 18:
            birth = datetime.strptime(id_str[6:14], '%Y%m%d').date()
        else:
            birth = datetime.strptime(id_str[6:12], '%y%m%d').date()
        age = get_age(birth, today)
        return age < 18
    except (ValueError, IndexError):
        return False

# ------------------------------
# 2. 文件读取引擎（xlrd 备选优化）
# ------------------------------

def read_calamine(filepath):
    """calamine 读取 .xlsx/.xls/.xlsm/.xlsb"""
    ids = set()
    try:
        wb = calamine_load(filepath)
        total_sheets = len(wb.sheet_names)
        print(f"  文件包含 {total_sheets} 个工作表")
        for idx, sheet_name in enumerate(wb.sheet_names, 1):
            sheet = wb.get_sheet_by_name(sheet_name)
            data = sheet.to_python()
            rows = len(data)
            print(f"  [{idx}/{total_sheets}] 工作表 '{sheet_name}' 共 {rows} 行，正在扫描...")
            for i, row in enumerate(data, 1):
                if i % 500 == 0 or i == rows:
                    print(f"    进度: {i}/{rows} 行", end='\r')
                for cell_value in row:
                    if cell_value is not None:
                        ids.update(extract_id_numbers(cell_value))
            print()
        wb.close()
    except Exception as e:
        print(f"  [错误] calamine 读取失败: {e}")
    return ids

def read_xls_by_xlrd(filepath):
    """xlrd 读取 .xls（兼容旧版格式，确保数字不转科学计数法）"""
    ids = set()
    xlrd_available = False
    try:
        import xlrd
        xlrd_available = True
    except ImportError:
        print("  [警告] 未安装 xlrd 库，无法使用备用引擎。请执行: pip install xlrd")
        return ids

    try:
        wb = xlrd.open_workbook(filepath)
        for ws in wb.sheets():
            print(f"  工作表 '{ws.name}' 共 {ws.nrows} 行，正在用 xlrd 扫描...")
            for r in range(ws.nrows):
                if r % 500 == 0:
                    print(f"    进度: {r}/{ws.nrows} 行", end='\r')
                for c in range(ws.ncols):
                    cell = ws.cell(r, c)
                    if cell.ctype == xlrd.XL_CELL_EMPTY:
                        continue
                    # 直接使用 cell.value，safe_str 会在 extract_id_numbers 中被调用
                    ids.update(extract_id_numbers(cell.value))
            print()
    except Exception as e:
        print(f"  [错误] xlrd 读取失败: {e}")
    return ids

def read_xls(filepath):
    """.xls 先用 calamine，若结果为空再用 xlrd（如果可用）"""
    ids = read_calamine(filepath)
    if not ids:
        print("  calamine 未提取到身份证号，尝试 xlrd 引擎...")
        ids = read_xls_by_xlrd(filepath)
    return ids

def read_et_by_wps(filepath):
    ids = set()
    try:
        import win32com.client
    except ImportError:
        print("  [提示] 需要 pywin32 库，尝试 olefile 后备...")
        return read_et_by_olefile(filepath)

    prog_ids = ["ET.Application", "KET.Application", "WPS.ET.Application"]
    et = None
    for prog in prog_ids:
        try:
            et = win32com.client.Dispatch(prog)
            print(f"  使用 WPS ProgID: {prog}")
            break
        except:
            continue

    if et is None:
        print("  未找到 WPS 组件，切换 olefile 后备方案")
        return read_et_by_olefile(filepath)

    try:
        et.Visible = False
        et.DisplayAlerts = False
        abs_path = os.path.abspath(filepath)
        wb = et.Workbooks.Open(abs_path, ReadOnly=True)
        time.sleep(1)
        for ws in wb.Worksheets:
            used = ws.UsedRange
            if used is None:
                continue
            rows = used.Rows.Count
            cols = used.Columns.Count
            print(f"  工作表 '{ws.Name}' 大小: {rows} 行 x {cols} 列，正在提取...")
            for r in range(1, rows + 1):
                if r % 200 == 0:
                    print(f"    进度: {r}/{rows} 行", end='\r')
                for c in range(1, cols + 1):
                    try:
                        val = ws.Cells(r, c).Value
                        if val is not None:
                            ids.update(extract_id_numbers(val))
                    except:
                        continue
            print()
        wb.Close(False)
    except Exception as e:
        print(f"  [错误] WPS 读取失败: {e}")
        print("  回退到 olefile 方案")
        ids = read_et_by_olefile(filepath)
    finally:
        try:
            et.Quit()
        except:
            pass
    return ids

def read_et_by_olefile(filepath):
    ids = set()
    try:
        if not olefile.isOleFile(filepath):
            return ids
        ole = olefile.OleFileIO(filepath)
        all_texts = []
        for stream in ole.listdir():
            try:
                data = ole.openstream(stream).read()
                for enc in ('utf-16-le', 'gbk', 'utf-8'):
                    try:
                        text = data.decode(enc, errors='ignore')
                        all_texts.append(text)
                        break
                    except:
                        continue
            except:
                continue
        ole.close()
        combined = ' '.join(all_texts)
        ids = extract_id_numbers(combined)
    except Exception as e:
        print(f"  [警告] olefile 读取失败: {e}")
    return ids

def read_et_file(filepath):
    ids = read_et_by_wps(filepath)
    if not ids:
        print("  WPS 未提取到数据，尝试 olefile...")
        ids = read_et_by_olefile(filepath)
    return ids

def read_excel_cells(filepath):
    ext = os.path.splitext(filepath)[1].lower()
    if ext == '.xls':
        return read_xls(filepath)
    elif ext in ['.xlsx', '.xlsm', '.xlsb']:
        return read_calamine(filepath)
    elif ext == '.et':
        return read_et_file(filepath)
    else:
        return set()

# ------------------------------
# 3. 主流程
# ------------------------------

def main():
    data_dir = 'data'
    if not os.path.isdir(data_dir):
        print(f"致命错误：文件夹 '{data_dir}' 不存在!")
        input("按回车键退出...")
        return

    supported_ext = ('.xlsx', '.xls', '.xlsm', '.xlsb', '.et')
    excel_files = []

    print(f"正在递归扫描 {data_dir} 中的表格文件...")
    for root, dirs, files in os.walk(data_dir):
        for f in files:
            if f.lower().endswith(supported_ext):
                full_path = os.path.join(root, f)
                excel_files.append(full_path)
                print(f"  找到: {full_path}")

    if not excel_files:
        print("未找到任何表格文件。")
        input("按回车键退出...")
        return

    all_ids = set()
    for filepath in excel_files:
        print(f"\n正在处理: {filepath}")
        try:
            ids = read_excel_cells(filepath)
            print(f"  该文件提取到 {len(ids)} 个有效身份证号")
            all_ids.update(ids)
        except KeyboardInterrupt:
            print("\n  您按下了 Ctrl+C，是否跳过此文件并继续？(y/n)")
            choice = input().strip().lower()
            if choice == 'n':
                print("  用户终止程序。")
                break
            else:
                print("  跳过该文件，继续处理下一个。")
                continue

    print(f"\n===== 汇总 =====")
    print(f"总共发现 {len(all_ids)} 个唯一的身份证号")

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "身份证号"
        ws.append(["身份证号"])

        red_font = Font(color="FF0000")
        today = date.today()

        if all_ids:
            for id_num in sorted(all_ids):
                ws.append([id_num])
                cell = ws.cell(row=ws.max_row, column=1)
                if is_minor(id_num, today):
                    cell.font = red_font
            print(f"未成年身份证号已标记为红色（截至 {today}）")
        else:
            print("未提取到任何身份证号，生成空表头。")

        output_file = "newTarget.xlsx"
        wb.save(output_file)
        abs_path = os.path.abspath(output_file)
        print(f"\n✅ 文件已保存到: {abs_path}")
    except Exception as e:
        print(f"\n❌ 保存失败: {e}")

    input("\n按回车键退出...")

if __name__ == "__main__":
    main()